from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import ChangeEvent, MatchResult, ProjectRow


PROJECTS_SHEET = "projects"
HISTORY_SHEET = "history"
MATCHES_SHEET = "matches"
MINENERGO_SHEET = "minenergo_3y"


PROJECT_COLUMNS = [
    "IDProject",
    "Title",
    "Link",
    "Stage",
    "Status",
    "PublishDate",
    "Date",
    "StartDiscussion",
    "EndDiscussion",
    "CreatorDepartment",
    "CreatorDepartmentReal",
    "Category",
    "Kind",
    "DegreeRegulatoryImpact",
    "Published",
    "RegionSignificant",
    "ControlSupervisoryActivities",
    "RegulatorScissors",
    "FirstSeenAt",
    "LastSeenAt",
    "LastChangedAt",
    "MatchedDocs",
    "MatchScore",
    "MatchExplain",
]

HISTORY_COLUMNS = ["ts", "IDProject", "field", "old", "new"]

MATCHES_COLUMNS = [
    "IDProject",
    "ProjectCardUrl",
    "ProjectDate",
    "ProjectInitiator",
    "Title",
    "Stage",
    "Status",
    "MatchScore",
    "MatchExplain",
    "MatchedDoc",
    "LastSeenAt",
]

MINENERGO_COLUMNS = [
    "IDProject",
    "ProjectCardUrl",
    "ProjectDate",
    "ProjectInitiator",
    "Title",
    "Stage",
    "Status",
    "PublishDate",
    "StartDiscussion",
    "EndDiscussion",
    "Category",
    "Kind",
]

HIGHLIGHT_COLORS = [
    "FFF2CC",  # light yellow
    "D9EAD3",  # light green
    "D9EAF7",  # light blue
    "FCE4D6",  # light orange
    "EADCF8",  # light purple
    "DDEBF7",
    "E2F0D9",
    "F4CCCC",
]


def _ensure_sheet(wb: Workbook, name: str, columns: list[str]) -> Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row < 1:
        for i, c in enumerate(columns, start=1):
            ws.cell(1, i).value = c
        return ws
    # Preserve existing data when adding new columns to a sheet schema.
    for i, c in enumerate(columns, start=1):
        if (ws.cell(1, i).value or "") == c:
            continue
        headers = [(ws.cell(1, j).value or "") for j in range(1, ws.max_column + 1)]
        if c not in headers:
            ws.insert_cols(i)
            ws.cell(1, i).value = c
    return ws


def _delete_rows_desc(ws: Worksheet, rows: list[int]) -> None:
    for r in sorted(rows, reverse=True):
        ws.delete_rows(r, 1)


def _delete_matches_rows_for_pid(ws: Worksheet, id_project: int) -> None:
    to_del = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == id_project]
    _delete_rows_desc(ws, to_del)


def _delete_indexed_row(ws: Worksheet, idx: dict[int, int], id_project: int) -> None:
    row_idx = idx.pop(id_project, None)
    if row_idx is None:
        return
    ws.delete_rows(row_idx, 1)
    for pid, existing_row in list(idx.items()):
        if existing_row > row_idx:
            idx[pid] = existing_row - 1


def _row_index_by_id(ws: Worksheet) -> dict[int, int]:
    # map IDProject -> row number
    idx: dict[int, int] = {}
    col_id = 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_id).value
        try:
            pid = int(v)
        except Exception:
            continue
        idx[pid] = r
    return idx


def _used_fill_colors(ws: Worksheet) -> set[str]:
    colors: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            fill = cell.fill
            if fill.fill_type != "solid":
                continue
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "00FFFFFF", "FFFFFFFF"):
                colors.add(rgb[-6:].upper())
    return colors


def _fill_for(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _apply_row_fill(ws: Worksheet, row_idx: int, color: str) -> None:
    fill = _fill_for(color)
    for cell in ws[row_idx]:
        cell.fill = fill


def _is_minenergo(row: ProjectRow) -> bool:
    haystack = f"{row.creator_department or ''} {row.creator_department_real or ''}".casefold()
    return "минэнерго" in haystack


def _dt_iso(dt: datetime) -> str:
    return dt.replace(microsecond=0).isoformat()


def _d_iso(d) -> str | None:
    return None if d is None else str(d)


def _b(v) -> str | None:
    if v is None:
        return None
    return "1" if bool(v) else "0"


class ExcelRegistry:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if self.path.is_file():
            self.wb = load_workbook(self.path)
        else:
            self.wb = Workbook()
        # openpyxl creates default sheet named "Sheet"
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) == 1:
            self.wb["Sheet"].title = PROJECTS_SHEET
        self.ws_projects = _ensure_sheet(self.wb, PROJECTS_SHEET, PROJECT_COLUMNS)
        self.ws_history = _ensure_sheet(self.wb, HISTORY_SHEET, HISTORY_COLUMNS)
        self.ws_matches = _ensure_sheet(self.wb, MATCHES_SHEET, MATCHES_COLUMNS)
        self.ws_minenergo = _ensure_sheet(self.wb, MINENERGO_SHEET, MINENERGO_COLUMNS)
        self._project_idx = _row_index_by_id(self.ws_projects)
        self._minenergo_idx = _row_index_by_id(self.ws_minenergo)

    def save(self) -> None:
        self.wb.save(self.path)

    def has_project(self, id_project: int) -> bool:
        return id_project in self._project_idx

    def next_highlight_color(self) -> str:
        used = _used_fill_colors(self.ws_projects) | _used_fill_colors(self.ws_matches)
        for color in HIGHLIGHT_COLORS:
            if color not in used:
                return color
        return HIGHLIGHT_COLORS[len(used) % len(HIGHLIGHT_COLORS)]

    def upsert(
        self,
        *,
        row: ProjectRow,
        match: MatchResult,
        first_seen_at: str,
        last_seen_at: str,
        last_changed_at: str | None,
        highlight_color: str | None = None,
    ) -> None:
        ws = self.ws_projects
        r = self._project_idx.get(row.id_project)
        if r is None:
            r = ws.max_row + 1
            self._project_idx[row.id_project] = r

        values = {
            "IDProject": row.id_project,
            "Title": row.title,
            "Link": row.link,
            "Stage": row.stage,
            "Status": row.status,
            "PublishDate": _d_iso(row.publish_date),
            "Date": _d_iso(row.date),
            "StartDiscussion": _d_iso(row.start_discussion),
            "EndDiscussion": _d_iso(row.end_discussion),
            "CreatorDepartment": row.creator_department,
            "CreatorDepartmentReal": row.creator_department_real,
            "Category": row.category,
            "Kind": row.kind,
            "DegreeRegulatoryImpact": row.degree_regulatory_impact,
            "Published": _b(row.published),
            "RegionSignificant": _b(row.region_significant),
            "ControlSupervisoryActivities": _b(row.control_supervisory_activities),
            "RegulatorScissors": _b(row.regulator_scissors),
            "FirstSeenAt": first_seen_at,
            "LastSeenAt": last_seen_at,
            "LastChangedAt": last_changed_at,
            "MatchedDocs": json.dumps(match.matched_docs, ensure_ascii=False),
            "MatchScore": match.score,
            "MatchExplain": match.explain,
        }

        for c_i, col in enumerate(PROJECT_COLUMNS, start=1):
            ws.cell(r, c_i).value = values.get(col)
        if highlight_color:
            _apply_row_fill(ws, r, highlight_color)

        self.sync_matches_sheet(
            row=row,
            match=match,
            last_seen_at=last_seen_at,
            highlight_color=highlight_color,
        )
        self.sync_minenergo_sheet(row=row, match=match, highlight_color=highlight_color)

    def sync_matches_sheet(
        self,
        *,
        row: ProjectRow,
        match: MatchResult,
        last_seen_at: str,
        highlight_color: str | None = None,
    ) -> None:
        """Одна строка на связку проект ↔ затронутый НПА."""
        wm = self.ws_matches
        _delete_matches_rows_for_pid(wm, row.id_project)
        if not match.matched_docs:
            return
        for md in match.matched_docs:
            rr = wm.max_row + 1
            wm.cell(rr, 1).value = row.id_project
            wm.cell(rr, 2).value = row.link
            if row.link:
                wm.cell(rr, 2).hyperlink = row.link
                wm.cell(rr, 2).style = "Hyperlink"
            wm.cell(rr, 3).value = _d_iso(row.date)
            wm.cell(rr, 4).value = row.creator_department
            wm.cell(rr, 5).value = row.title
            wm.cell(rr, 6).value = row.stage
            wm.cell(rr, 7).value = row.status
            wm.cell(rr, 8).value = match.score
            wm.cell(rr, 9).value = match.explain
            wm.cell(rr, 10).value = md
            wm.cell(rr, 11).value = last_seen_at
            if highlight_color:
                _apply_row_fill(wm, rr, highlight_color)

    def sync_minenergo_sheet(
        self,
        *,
        row: ProjectRow,
        match: MatchResult,
        highlight_color: str | None = None,
    ) -> None:
        if not _is_minenergo(row) or match.matched_docs:
            _delete_indexed_row(self.ws_minenergo, self._minenergo_idx, row.id_project)
            return

        ws = self.ws_minenergo
        r = self._minenergo_idx.get(row.id_project)
        if r is None:
            r = ws.max_row + 1
            self._minenergo_idx[row.id_project] = r

        values = [
            row.id_project,
            row.link,
            _d_iso(row.date),
            row.creator_department,
            row.title,
            row.stage,
            row.status,
            _d_iso(row.publish_date),
            _d_iso(row.start_discussion),
            _d_iso(row.end_discussion),
            row.category,
            row.kind,
        ]
        for c_i, value in enumerate(values, start=1):
            ws.cell(r, c_i).value = value
        if row.link:
            ws.cell(r, 2).hyperlink = row.link
            ws.cell(r, 2).style = "Hyperlink"
        if highlight_color:
            _apply_row_fill(ws, r, highlight_color)

    def append_history(self, events: list[ChangeEvent]) -> None:
        ws = self.ws_history
        for ev in events:
            r = ws.max_row + 1
            ws.cell(r, 1).value = _dt_iso(ev.ts)
            ws.cell(r, 2).value = ev.id_project
            ws.cell(r, 3).value = ev.field
            ws.cell(r, 4).value = ev.old
            ws.cell(r, 5).value = ev.new

